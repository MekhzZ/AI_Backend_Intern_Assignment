import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

# Load the Excel files
assignment_data_file = "./assets/excelFile/Assignment Data.xlsx"
video_predictions_file = "./outputFiles/video_predictions.xlsx"

# Read the data
assignment_data = pd.read_excel(assignment_data_file)
video_predictions = pd.read_excel(video_predictions_file)

# Extract matching identifiers
assignment_data['Identifier'] = assignment_data['Video URL'].apply(lambda x: x.split('/')[-1])
video_predictions['Identifier'] = video_predictions['Video File'].apply(lambda x: x.replace('.mp4', ''))

# Merge the data based on the extracted identifier
merged_data = pd.merge(
    assignment_data,
    video_predictions,
    on="Identifier",
    how="inner"
)

# Save the merged data to a temporary Excel file
output_file = "merged_video_data_with_images.xlsx"
merged_data.to_excel(output_file, index=False)

# Open the workbook for adding images
wb = load_workbook(output_file)
ws = wb.active

# Folder containing pre-downloaded images
image_folder = "./pre_downloaded_images"

# Map labels to their corresponding image files
label_to_image = {
    "messi": os.path.join(image_folder, "messi.jpg"),  # Replace with actual file names
    "ronaldo": os.path.join(image_folder, "ronaldo.jpg"),
    "diljith": os.path.join(image_folder, "diljith.jpg"),
}

# Add images to the Excel sheet
for row in range(2, ws.max_row + 1):  # Start from row 2 to skip the header
    predicted_label = ws[f'E{row}'].value  # Adjust 'E' to the column with "Predicted Label"
    if not predicted_label:
        continue
    
    # Select only the first word of the predicted label
    predicted_label_first_word = predicted_label.split(',')[0]
    
    if predicted_label_first_word not in label_to_image:
        continue

    # Get the image path for the predicted label
    image_path = label_to_image[predicted_label_first_word]

    # Insert the image into the Excel sheet
    try:
        if os.path.exists(image_path):
            img = Image(image_path)
            img.width = 100  # Adjust width (optional)
            img.height = 100  # Adjust height (optional)
            ws.add_image(img, f'F{row}')  # Add image to the 'F' column
        else:
            print(f"Image not found for {predicted_label_first_word}")
    except Exception as e:
        print(f"Error inserting image for {predicted_label_first_word}: {e}")

# Auto-fit row and column dimensions
def autofit(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.rows:
        max_height = 0
        for cell in row:
            try:
                if len(str(cell.value)) > max_height:
                    max_height = len(cell.value)
            except:
                pass
        adjusted_height = max_height + 5
        ws.row_dimensions[row[0].row].height = adjusted_height

# Apply auto-fit
autofit(ws)

# Save the workbook
wb.save(output_file)
print(f"Merged file with images saved as {output_file}")
